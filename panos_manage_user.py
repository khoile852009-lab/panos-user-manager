#!/usr/bin/env python3
"""
panos_manage_user.py - Manage local admin accounts on PAN-OS firewalls or Panoramas.

Usage:
    python panos_manage_user.py --firewall hosts.txt --auth admin:password --operation list
    python panos_manage_user.py --panorama hosts.txt --auth-api-key APIKEY --operation create --username bob --role superuser
    python panos_manage_user.py --firewall hosts.txt --auth admin:password --operation update --username bob
    python panos_manage_user.py --panorama hosts.txt --auth admin:password --operation delete --username bob

Supported roles:
    superuser, superreader, device-admin, device-admin-read-only, or any custom admin role profile name
"""

import argparse
import sys
import logging
import os
import time
import getpass
import xml.etree.ElementTree as ET
import base64
from typing import Optional
from urllib.parse import urlencode
from urllib.request import urlopen, Request
from urllib.error import URLError, HTTPError

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logging.basicConfig(level=logging.WARNING, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

ADMIN_XPATH = "/config/mgt-config/users"
TEMPLATE_BASE_XPATH = "/config/devices/entry[@name='localhost.localdomain']/template"
W = 68  # output width


def banner(text: str) -> None:
    print("=" * W)
    print(f"  {text}")
    print("=" * W)


def section(host: str, index: int, total: int) -> None:
    print(f"\n[{index}/{total}] {host}")
    print("-" * W)


def ok(msg: str) -> None:
    print(f"  [+] {msg}")


def fail(msg: str) -> None:
    print(f"  [-] {msg}")


def info(msg: str) -> None:
    print(f"  ... {msg}")


def resolve_auth(args) -> tuple:
    """Return (headers dict, commit_admin username or None).

    Priority order for each credential:
      1. CLI arg   2. Environment variable   3. Interactive prompt
    """
    # API key path
    api_key = args.auth_api_key or os.environ.get("PANOS_API_KEY")
    if api_key:
        return {"X-PAN-KEY": api_key}, None

    # Username / password path
    auth_str = args.auth or os.environ.get("PANOS_AUTH", "")
    if ":" in auth_str:
        username, password = auth_str.split(":", 1)
    else:
        username = auth_str or input("Login username: ").strip()
        password = os.environ.get("PANOS_PASSWORD") or getpass.getpass(f"Login password for '{username}': ")

    token = base64.b64encode(f"{username}:{password}".encode()).decode()
    return {"Authorization": f"Basic {token}"}, username


def resolve_user_password(args) -> str:
    """Return the target account password from CLI arg, env var, or prompt."""
    if args.password:
        return args.password
    env_pw = os.environ.get("PANOS_USER_PASSWORD")
    if env_pw:
        return env_pw
    return getpass.getpass(f"Password for new/updated user '{args.username}': ")


def api_request(host: str, params: dict, headers: dict) -> ET.Element:
    url = f"https://{host}/api/?" + urlencode(params)
    req = Request(url, headers=headers)
    try:
        with urlopen(req) as resp:
            body = resp.read()
    except HTTPError as e:
        body = e.read()
        raise RuntimeError(f"HTTP {e.code}: {e.reason} — {body.decode(errors='replace')}") from e
    except URLError as e:
        raise RuntimeError(f"Connection error: {e.reason}") from e

    root = ET.fromstring(body)
    status = root.get("status")
    if status != "success":
        msg = root.findtext(".//msg") or ET.tostring(root, encoding="unicode")
        raise RuntimeError(f"API error ({status}): {msg}")
    return root


def hash_password(host: str, headers: dict, password: str) -> str:
    cmd_root = ET.Element("request")
    pw_hash = ET.SubElement(cmd_root, "password-hash")
    pw = ET.SubElement(pw_hash, "password")
    pw.text = password
    cmd_xml = ET.tostring(cmd_root, encoding="unicode")

    root = api_request(host, {"type": "op", "cmd": cmd_xml}, headers)
    phash = root.findtext(".//phash")
    if not phash:
        raise RuntimeError("password-hash response missing <phash>")
    return phash


def _parse_role(entry: ET.Element) -> str:
    role_based = entry.find("permissions/role-based")
    if role_based is None:
        return "unknown"
    if role_based.find("superuser") is not None:
        return "superuser"
    if role_based.find("superreader") is not None:
        return "superreader"
    if role_based.find("deviceadmin") is not None:
        return "device-admin"
    if role_based.find("devicereader") is not None:
        return "device-admin-read-only"
    custom = role_based.find("custom/profile")
    if custom is not None and custom.text:
        return f"custom:{custom.text}"
    return "unknown"


def _build_role_element(role: str) -> ET.Element:
    permissions = ET.Element("permissions")
    role_based = ET.SubElement(permissions, "role-based")
    role_lower = role.lower()

    if role_lower == "superuser":
        ET.SubElement(role_based, "superuser").text = "yes"
    elif role_lower == "superreader":
        ET.SubElement(role_based, "superreader").text = "yes"
    elif role_lower in ("device-admin", "deviceadmin"):
        ET.SubElement(role_based, "deviceadmin").text = "yes"
    elif role_lower in ("device-admin-read-only", "deviceadminreadonly"):
        ET.SubElement(role_based, "devicereader").text = "yes"
    else:
        custom = ET.SubElement(role_based, "custom")
        profile = ET.SubElement(custom, "profile")
        profile.text = role

    return permissions


# ── template helpers ─────────────────────────────────────────────────────────

def template_admin_xpath(template_name: str) -> str:
    """Return the mgt-config/users XPath inside a specific device template."""
    return (
        f"{TEMPLATE_BASE_XPATH}/entry[@name='{template_name}']"
        "/config/mgt-config/users"
    )


def get_template_names(host: str, headers: dict) -> list:
    """Return all device-template names from a Panorama."""
    try:
        root = api_request(host, {
            "type": "config",
            "action": "get",
            "xpath": TEMPLATE_BASE_XPATH,
        }, headers)
    except RuntimeError as e:
        fail(f"Could not retrieve device templates: {e}")
        return []
    # Only grab direct <entry> children of <template>, not nested entries
    template_node = root.find(".//template")
    if template_node is None:
        # API may return <result><template><entry ...>... so try result/entry
        template_node = root.find(".//result")
    if template_node is None:
        return []
    return [e.get("name") for e in template_node.findall("entry") if e.get("name")]


# ── operations ───────────────────────────────────────────────────────────────

def op_list(host: str, headers: dict, xpath: str = ADMIN_XPATH, label: str = "") -> bool:
    if label:
        info(label)
    try:
        root = api_request(host, {
            "type": "config",
            "action": "get",
            "xpath": xpath,
        }, headers)
    except RuntimeError as e:
        if "No such node" in str(e) or "Object not found" in str(e):
            info("  No local admins" if label else "No local admins found")
            return True
        fail(f"Could not retrieve users: {e}")
        return False

    entries = root.findall(".//entry")
    if not entries:
        info("  No local admins" if label else "No local admins found")
        return True

    for entry in entries:
        name = entry.get("name", "?")
        role = _parse_role(entry)
        print(f"  {'':2}{name:<30} role={role}")
    return True


def collect_list(host: str, headers: dict, xpath: str = ADMIN_XPATH,
                 label: str = "") -> list:
    """Return a list of (scope_label, username, role) tuples for xlsx output."""
    rows = []
    try:
        root = api_request(host, {
            "type": "config",
            "action": "get",
            "xpath": xpath,
        }, headers)
    except RuntimeError:
        return rows

    for entry in root.findall(".//entry"):
        name = entry.get("name", "?")
        role = _parse_role(entry)
        rows.append((label or "Panorama-level", name, role))
    return rows


def write_xlsx(path: str, host_data: dict) -> None:
    """Write collected list data to an xlsx file (one sheet per host)."""
    wb = Workbook()
    # Remove the default sheet created by openpyxl
    wb.remove(wb.active)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for host, rows in host_data.items():
        # Sheet names limited to 31 chars and certain chars are invalid
        sheet_name = host[:31].replace("/", "_").replace("\\", "_").replace(":", "_")
        ws = wb.create_sheet(title=sheet_name)

        headers_row = ["Scope", "Username", "Role"]
        ws.append(headers_row)
        for col_idx in range(1, len(headers_row) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")

        if rows:
            for scope_label, username, role in rows:
                ws.append([scope_label, username, role])
        else:
            ws.append(["(no local admins found)", "", ""])

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = max_len + 3

    wb.save(path)


def op_create(host: str, headers: dict, username: str, password: str, role: str,
              xpath: str = ADMIN_XPATH) -> bool:
    try:
        info("Hashing password...")
        phash = hash_password(host, headers, password)

        entry = ET.Element("entry", attrib={"name": username})
        phash_el = ET.SubElement(entry, "phash")
        phash_el.text = phash
        entry.append(_build_role_element(role))

        api_request(host, {
            "type": "config",
            "action": "set",
            "xpath": xpath,
            "element": ET.tostring(entry, encoding="unicode"),
        }, headers)
        ok(f"User '{username}' created  (role={role})")
        return True
    except RuntimeError as e:
        fail(f"Create '{username}' failed: {e}")
        return False


def op_update(host: str, headers: dict, username: str,
              password: Optional[str], role: Optional[str],
              xpath: str = ADMIN_XPATH) -> bool:
    if not password and not role:
        info("Nothing to update — pass --password and/or --role")
        return True

    xpath_entry = f"{xpath}/entry[@name='{username}']"

    try:
        api_request(host, {"type": "config", "action": "get", "xpath": xpath_entry}, headers)
    except RuntimeError:
        fail(f"User '{username}' not found")
        return False

    changed = []
    try:
        if password:
            info("Hashing password...")
            phash = hash_password(host, headers, password)
            phash_el = ET.Element("phash")
            phash_el.text = phash
            api_request(host, {
                "type": "config",
                "action": "edit",
                "xpath": f"{xpath_entry}/phash",
                "element": ET.tostring(phash_el, encoding="unicode"),
            }, headers)
            changed.append("password")

        if role:
            api_request(host, {
                "type": "config",
                "action": "edit",
                "xpath": f"{xpath_entry}/permissions",
                "element": ET.tostring(_build_role_element(role), encoding="unicode"),
            }, headers)
            changed.append(f"role={role}")

        ok(f"User '{username}' updated  ({', '.join(changed)})")
        return True
    except RuntimeError as e:
        fail(f"Update '{username}' failed: {e}")
        return False


def op_delete(host: str, headers: dict, username: str, xpath: str = ADMIN_XPATH) -> bool:
    xpath_entry = f"{xpath}/entry[@name='{username}']"
    try:
        api_request(host, {
            "type": "config",
            "action": "delete",
            "xpath": xpath_entry,
        }, headers)
        ok(f"User '{username}' deleted")
        return True
    except RuntimeError as e:
        if "not found" in str(e).lower() or "No such node" in str(e):
            fail(f"User '{username}' not found")
        else:
            fail(f"Delete '{username}' failed: {e}")
        return False


def do_commit(host: str, headers: dict, admin_user: Optional[str]) -> bool:
    commit = ET.Element("commit")
    if admin_user:
        partial = ET.SubElement(commit, "partial")
        admins = ET.SubElement(partial, "admin")
        member = ET.SubElement(admins, "member")
        member.text = admin_user

    scope = f"partial (admin={admin_user})" if admin_user else "full"

    try:
        resp = api_request(host, {
            "type": "commit",
            "cmd": ET.tostring(commit, encoding="unicode"),
        }, headers)
    except RuntimeError as e:
        fail(f"Commit failed: {e}")
        return False

    job_id = resp.findtext(".//job")
    if not job_id:
        msg = resp.findtext(".//msg") or resp.findtext(".//line") or "no pending changes"
        info(f"Commit ({scope}): {msg}")
        return True

    info(f"Commit job {job_id} queued ({scope}) — polling...")

    show_cmd = ET.tostring(
        ET.fromstring(f"<show><jobs><id>{job_id}</id></jobs></show>"),
        encoding="unicode",
    )

    for elapsed in range(60):
        time.sleep(3)
        try:
            poll = api_request(host, {"type": "op", "cmd": show_cmd}, headers)
        except RuntimeError as e:
            fail(f"Poll job {job_id} failed: {e}")
            return False

        status = poll.findtext(".//job/status") or ""
        if status != "FIN":
            if elapsed % 5 == 4:
                info(f"  Job {job_id} still running ({(elapsed + 1) * 3}s)...")
            continue

        result = poll.findtext(".//job/result") or "unknown"
        details = [el.text for el in poll.findall(".//job/details/line") if el.text]
        detail_str = " | ".join(details) if details else ""
        suffix = f": {detail_str}" if detail_str else ""

        if result == "OK":
            ok(f"Commit job {job_id} succeeded{suffix}")
            return True
        else:
            fail(f"Commit job {job_id} failed ({result}){suffix}")
            return False

    fail(f"Commit job {job_id} still running after 3 min — check device manually")
    return False


# ── helpers ───────────────────────────────────────────────────────────────────

def load_hosts(path: str) -> list:
    try:
        with open(path) as f:
            return [line.strip() for line in f if line.strip() and not line.startswith("#")]
    except FileNotFoundError:
        print(f"ERROR: host file not found: {path}")
        sys.exit(1)


def print_summary(results: list, operation: str) -> None:
    print()
    banner("Summary")
    total = len(results)
    op_ok = sum(1 for r in results if r["op_ok"])
    op_fail = total - op_ok

    col = 22
    print(f"  {'Operation':<{col}} {operation.upper()}")
    print(f"  {'Hosts processed':<{col}} {total}")
    print(f"  {'Succeeded':<{col}} {op_ok}")
    if op_fail:
        print(f"  {'Failed':<{col}} {op_fail}")

    commit_results = [r for r in results if r.get("commit_ok") is not None]
    if commit_results:
        c_ok = sum(1 for r in commit_results if r["commit_ok"])
        c_fail = len(commit_results) - c_ok
        print(f"  {'Commits OK':<{col}} {c_ok}")
        if c_fail:
            print(f"  {'Commits failed':<{col}} {c_fail}")

    if op_fail or any(not r.get("commit_ok", True) for r in commit_results):
        print()
        print("  Failed hosts:")
        for r in results:
            issues = []
            if not r["op_ok"]:
                issues.append("operation")
            if r.get("commit_ok") is False:
                issues.append("commit")
            if issues:
                print(f"    - {r['host']}  ({', '.join(issues)})")

    print("=" * W)


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Manage local admin accounts on PAN-OS firewalls or Panoramas",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )

    target = parser.add_mutually_exclusive_group(required=True)
    target.add_argument("--firewall", metavar="FILE", help="File of firewall hosts")
    target.add_argument("--panorama", metavar="FILE", help="File of Panorama hosts")

    auth = parser.add_mutually_exclusive_group(required=False)
    auth.add_argument("--auth", metavar="USER[:PASS]",
                      help="Login as USER:PASS (password prompted if omitted; "
                           "or set PANOS_AUTH / PANOS_PASSWORD env vars)")
    auth.add_argument("--auth-api-key", metavar="KEY",
                      help="API key sent as X-PAN-KEY header (or set PANOS_API_KEY env var)")

    parser.add_argument("--operation", required=True,
                        choices=["list", "create", "update", "delete"])
    parser.add_argument("--username", help="Target username (required for create/update/delete)")
    parser.add_argument("--password", help="Password (required for create, optional for update)")
    parser.add_argument("--role", default=None,
                        help="Role: superuser (default for create), superreader, device-admin, "
                             "device-admin-read-only, or a custom admin role profile name")
    parser.add_argument("--commit", action="store_true",
                        help="Partial-commit after each change using the login admin from --auth "
                             "(full commit when using --auth-api-key)")
    parser.add_argument("--output-xlsx", metavar="FILE",
                        help="Write list results to an xlsx file (one sheet per host). "
                             "Requires openpyxl: pip install openpyxl")
    parser.add_argument("--scope", choices=["panorama", "templates", "all"], default="panorama",
                        help="Scope: panorama (default), templates (device templates only), "
                             "or all (panorama-level + all device templates). "
                             "templates/all require --panorama.")
    parser.add_argument("--template", metavar="NAME",
                        help="Target a single device template by name "
                             "(used with --scope templates|all; omit to target all templates)")

    args = parser.parse_args()

    if args.operation in ("create", "update", "delete") and not args.username:
        parser.error(f"--username is required for {args.operation}")

    device_type = "panorama" if args.panorama else "firewall"

    if args.scope != "panorama" and device_type != "panorama":
        parser.error("--scope templates/all is only valid with --panorama")
    if args.output_xlsx and args.operation != "list":
        parser.error("--output-xlsx is only valid with --operation list")
    if args.output_xlsx and not HAS_OPENPYXL:
        print("ERROR: openpyxl is required for xlsx output.  pip install openpyxl")
        sys.exit(1)
    host_file = args.panorama or args.firewall
    hosts = load_hosts(host_file)
    if not hosts:
        print("ERROR: no hosts found in file")
        sys.exit(1)

    # Resolve credentials — may prompt interactively
    headers, commit_admin = resolve_auth(args)

    # Resolve target user password if needed — may prompt interactively
    user_password = ""
    if args.operation == "create":
        user_password = resolve_user_password(args)
    elif args.operation == "update":
        # Password is optional for update; only prompt when explicitly needed
        user_password = args.password or os.environ.get("PANOS_USER_PASSWORD", "")

    col = 22
    banner("PAN-OS User Manager")
    print(f"  {'Operation':<{col}} {args.operation.upper()}")
    print(f"  {'Device type':<{col}} {device_type}")
    if args.scope != "panorama":
        scope_label = args.template or "all templates"
        print(f"  {'Scope':<{col}} {args.scope}  ({scope_label})")
    print(f"  {'Hosts':<{col}} {len(hosts)}")
    if args.username:
        print(f"  {'Target user':<{col}} {args.username}")
    effective_role = args.role or ("superuser" if args.operation == "create" else None)
    if effective_role:
        print(f"  {'Role':<{col}} {effective_role}")
    if args.commit:
        commit_scope = "partial" if commit_admin else "full"
        print(f"  {'Commit':<{col}} {commit_scope}")
    print("=" * W)

    results = []
    xlsx_data: dict = {}  # host -> [(scope, user, role), ...]

    for i, host in enumerate(hosts, 1):
        section(host, i, len(hosts))
        op_ok = True
        commit_ok = None

        # Build (xpath, label) pairs based on --scope / --template
        scope_targets: list = []
        if args.scope in ("panorama", "all"):
            scope_targets.append((ADMIN_XPATH, "Panorama-level"))
        if device_type == "panorama" and args.scope in ("templates", "all"):
            if args.template:
                scope_targets.append((template_admin_xpath(args.template), f"Template: {args.template}"))
            else:
                template_names = get_template_names(host, headers)
                if not template_names:
                    info("No device templates found")
                for t in template_names:
                    scope_targets.append((template_admin_xpath(t), f"Template: {t}"))

        host_rows: list = []
        multi = len(scope_targets) > 1
        for xpath, label in scope_targets:
            scope_label = label if multi else ""
            if args.operation == "list":
                scope_ok = op_list(host, headers, xpath, scope_label)
                if args.output_xlsx:
                    host_rows.extend(collect_list(host, headers, xpath, scope_label))
            elif args.operation == "create":
                if multi:
                    info(label)
                scope_ok = op_create(host, headers, args.username, user_password, args.role or "superuser", xpath)
            elif args.operation == "update":
                if multi:
                    info(label)
                scope_ok = op_update(host, headers, args.username, user_password or None,
                                     args.role, xpath)
            elif args.operation == "delete":
                if multi:
                    info(label)
                scope_ok = op_delete(host, headers, args.username, xpath)
            else:
                scope_ok = True
            op_ok = op_ok and scope_ok

        if args.output_xlsx:
            xlsx_data[host] = host_rows

        if args.commit and args.operation != "list":
            commit_ok = do_commit(host, headers, commit_admin)

        results.append({"host": host, "op_ok": op_ok, "commit_ok": commit_ok})

    if args.output_xlsx:
        write_xlsx(args.output_xlsx, xlsx_data)
        ok(f"Wrote {len(xlsx_data)} sheet(s) to {args.output_xlsx}")

    if args.operation != "list":
        print_summary(results, args.operation)


if __name__ == "__main__":
    main()
